import customtkinter
import os
import jinja2
import pdfkit
import win32com.client as win32
import shutil
from datetime import timedelta
from datetime import date
from os.path import expanduser
from tkinter import *
from typing import Union, Callable

import openpyxl

from CTkMessagebox import CTkMessagebox
from CTkPDFViewer import *

customtkinter.set_appearance_mode("dark")  # Modes: system (default), light, dark
customtkinter.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green


class WidgetName(customtkinter.CTkFrame):
    def __init__(self, *args,
                 width: int = 100,
                 height: int = 32,
                 **kwargs):
        super().__init__(*args, width=width, height=height, **kwargs)


class FloatSpinbox(customtkinter.CTkFrame):
    def __init__(self, *args,
                 width: int = 100,
                 height: int = 32,
                 step_size: Union[int, float] = 1,
                 labelvalue: str = "",
                 command: Callable = None,
                 **kwargs):
        super().__init__(*args, width=width, height=height, **kwargs)

        self.step_size = step_size
        self.command = command
        self.labelvalue = labelvalue
        self.configure(fg_color=("gray78", "gray28"))  # set frame color

        self.grid_columnconfigure((0, 2), weight=0)  # buttons don't expand
        self.grid_columnconfigure(1, weight=1)  # entry expands

        self.label = customtkinter.CTkLabel(self, text=self.labelvalue)
        self.label.grid(row=0, column=0, padx=(3, 0), pady=3, columnspan=3)
        self.subtract_button = customtkinter.CTkButton(self, text="-", width=height - 6, height=height - 6,
                                                       command=self.subtract_button_callback)
        self.subtract_button.grid(row=1, column=0, padx=(3, 0), pady=3)

        self.entry = customtkinter.CTkEntry(self, width=width - (2 * height), height=height - 6, border_width=0)
        self.entry.grid(row=1, column=1, columnspan=1, padx=3, pady=3, sticky="ew")

        self.add_button = customtkinter.CTkButton(self, text="+", width=height - 6, height=height - 6,
                                                  command=self.add_button_callback)
        self.add_button.grid(row=1, column=2, padx=(0, 3), pady=3)

        # default value
        self.entry.insert(0, "£0.00")

    def add_button_callback(self):
        if self.command is not None:
            self.command()
        try:
            string = self.entry.get()
            value = float(string[1:len(string)]) + self.step_size
            self.entry.delete(1, "end")
            self.entry.insert(1, value)
        except ValueError:
            return

    def subtract_button_callback(self):
        if self.command is not None:
            self.command()
        try:
            string = self.entry.get()
            value = float(string[1:len(string)]) - self.step_size
            self.entry.delete(1, "end")
            self.entry.insert(1, value)
        except ValueError:
            return

    def get(self) -> Union[float, None]:
        try:
            string = self.entry.get()
            if string[0] == "£":
                return float(string[1:len(string)])
            else:
                return float(string)
        except ValueError:
            return None

    def set(self, value: float):
        self.entry.delete(1, "end")
        self.entry.insert(1, str(float(value)))


class IntSpinbox(customtkinter.CTkFrame):
    def __init__(self, *args,
                 width: int = 100,
                 height: int = 32,
                 step_size: Union[int, float] = 1,
                 command: Callable = None,
                 **kwargs):
        super().__init__(*args, width=width, height=height, **kwargs)

        self.step_size = step_size
        self.command = command

        self.configure(fg_color=("gray78", "gray28"))  # set frame color

        self.grid_columnconfigure((0, 2), weight=0)  # buttons don't expand
        self.grid_columnconfigure(1, weight=1)  # entry expands

        self.subtract_button = customtkinter.CTkButton(self, text="-", width=height - 6, height=height - 6,
                                                       command=self.subtract_button_callback)
        self.subtract_button.grid(row=0, column=0, padx=(3, 0), pady=3)

        self.entry = customtkinter.CTkEntry(self, width=width - (2 * height), height=height - 6, border_width=0)
        self.entry.grid(row=0, column=1, columnspan=1, padx=3, pady=3, sticky="ew")

        self.add_button = customtkinter.CTkButton(self, text="+", width=height - 6, height=height - 6,
                                                  command=self.add_button_callback)
        self.add_button.grid(row=0, column=2, padx=(0, 3), pady=3)

        # default value
        self.entry.insert(0, "0")

    def add_button_callback(self):
        if self.command is not None:
            self.command()
        try:
            value = int(self.entry.get()) + self.step_size
            self.entry.delete(0, "end")
            self.entry.insert(0, value)
        except ValueError:
            return

    def subtract_button_callback(self):
        if self.command is not None:
            self.command()
        try:
            value = int(self.entry.get()) - self.step_size
            self.entry.delete(0, "end")
            self.entry.insert(0, value)
        except ValueError:
            return

    def get(self) -> Union[int, None]:
        try:
            return int(self.entry.get())
        except ValueError:
            return None

    def set(self, value: int):
        self.entry.delete(0, "end")
        self.entry.insert(0, str(int(value)))


class scrollableTextFrame(customtkinter.CTkFrame):
    def __init__(self, master):
        super().__init__(master)

        self.label = customtkinter.CTkLabel(self, text="Additional Info")
        self.label.pack()

        self.textbox = customtkinter.CTkTextbox(self)
        self.textbox.pack(fill=BOTH, expand=1)


class MyScrollableInputFrame(customtkinter.CTkScrollableFrame):
    def __init__(self, master, title, values, controller, wrap=None):
        super().__init__(master, label_text=title)
        self.controller = controller
        self.grid_columnconfigure(0, weight=1)
        self.values = values
        self.inputs = []
        self.textwrap = wrap
        for i, value in enumerate(self.values):
            entry = customtkinter.CTkEntry(self, placeholder_text=value)
            entry.grid(row=i, column=0, padx=10, pady=(10, 0), sticky="nsew")
            self.inputs.append(entry)

    def get(self):
        checked_inputs = []
        for input in self.inputs:
            if input.get() == 1:
                checked_inputs.append(input.cget("text"))
        return checked_inputs


class InvoiceInputFrame(customtkinter.CTkScrollableFrame):
    def __init__(self, master, title, values, controller):
        super().__init__(master, label_text=title)
        self.controller = controller
        self.grid_columnconfigure(0, weight=1)
        self.values = values
        self.inputs = []

        for i, value in enumerate(self.values):
            entry = FloatSpinbox(self, width=300, step_size=0.50, labelvalue=value)
            entry.grid(row=i, column=0, padx=10, pady=(10, 0), sticky="ew")
            self.inputs.append(entry)

    def get(self):
        checked_inputs = []
        for input in self.inputs:
            if input.get() == 1:
                checked_inputs.append(input.cget("text"))
        return checked_inputs


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        if not os.path.exists(expanduser('~/Documents/Invoices')):
            os.makedirs(expanduser('~/Documents/Invoices'))
            print('Folder not found so new folder created')
        else:
            print('Folder Found')
        self.title("Invoice Creator")
        self.geometry("1080x640")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.frames = []

        values = ["Recipient Name", "Recipient Email", "Recipient Phone Number", "Work Location"]
        self.frames.append(MyScrollableInputFrame(self, title="Recipient", values=values, controller=self))
        self.frames[0].grid(row=0, column=0, padx=10, pady=(10, 0), sticky="nsew")

        invoiceinputs = ["Travel", "Labour", "Materials", "Overnight"]
        self.frames.append(InvoiceInputFrame(self, title="Costs", values=invoiceinputs, controller=self))

        values = ["Description of work done (or other required info)"]
        self.frames.append(scrollableTextFrame(self))

        self.frames.append(CTkPDFViewer(self, file="preview_pdf.pdf"))

        self.nextbutton = customtkinter.CTkButton(self, text="Next", command=self.nextbutton_callback)
        self.nextbutton.grid(row=3, column=0, padx=10, pady=10, sticky="ew", columnspan=2)

        self.previousbutton = customtkinter.CTkButton(self, text="Previous", command=self.previousbutton_callback)
        self.previousbutton.grid(row=4, column=0, padx=10, pady=10, sticky="ew", columnspan=2)

        self.dest_dir = None
        self.src_file = None

        self.dst_file = None
        self.new_dst_file_name = None

    global template_loader
    template_loader = jinja2.FileSystemLoader('./')
    global template_env
    template_env = jinja2.Environment(loader=template_loader)
    global template
    template = template_env.get_template("invoice.html")

    def copy_same_file_pass(self, src, dst):
        try:
            shutil.copy(src,dst)
        except shutil.SameFileError:
            pass

    def nextbutton_callback(self):
        if self.frames[0].winfo_ismapped():
            self.frames[0].grid_forget()
            self.frames[1].grid(row=0, column=0, padx=10, pady=(10, 0), sticky="nsew")
        elif self.frames[1].winfo_ismapped():
            self.frames[1].grid_forget()
            self.frames[2].grid(row=0, column=0, padx=10, pady=(10, 0), sticky="nsew")
        elif self.frames[2].winfo_ismapped():
            today_date = date.today()
            new_today_date = today_date.strftime("%d/%m/%Y")
            due_date = today_date + timedelta(days=30)
            new_due_date = due_date.strftime("%d/%m/%Y")
            total_cost = self.frames[1].inputs[0].get() + self.frames[1].inputs[1].get() + self.frames[1].inputs[
                2].get() + self.frames[1].inputs[3].get()
            wb = openpyxl.load_workbook(expanduser('~/Documents/Invoices/invoice_list.xlsx'))
            ws = wb.active
            context = {'invoice_number': ws.max_row, 'recipient': self.frames[0].inputs[0].get(),
                       'recipient_email': self.frames[0].inputs[1].get(),
                       'recipient_phonenum': self.frames[0].inputs[2].get(),
                       'Work_Location': self.frames[0].inputs[3].get(), 'todays_date': new_today_date,
                       'due_date': new_due_date, 'travel_cost': str(self.frames[1].inputs[0].get()) + "0",
                       'labour_cost': str(self.frames[1].inputs[1].get()) + "0",
                       'materials_cost': str(self.frames[1].inputs[2].get()) + "0",
                       'overnight_cost': str(self.frames[1].inputs[3].get()) + "0", 'total_cost': str(total_cost) + "0",
                       'additional_info': self.frames[2].textbox.get(1.0, "end-1c")}

            output_text = template.render(context)

            config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
            pdfkit.from_string(output_text, 'preview_pdf.pdf', configuration=config)

            self.frames[2].grid_forget()
            self.frames[3].grid(row=0, column=0, padx=10, pady=(10, 0), sticky="nsew")
            self.frames[3].configure(file='preview_pdf.pdf')
        elif self.frames[3].winfo_ismapped():
            msg = CTkMessagebox(title="Is PDF Correct", message="Is the PDF correct and ready to be sent?",
                                icon="question", option_1="No", option_2="Yes")
            response = msg.get()

            if response == "Yes":
                wb = openpyxl.load_workbook(expanduser('~/Documents/Invoices/invoice_list.xlsx'))
                ws = wb.active
                today_date = date.today()
                new_today_date = today_date.strftime("%d/%m/%Y")
                due_date = today_date + timedelta(days=30)
                new_due_date = due_date.strftime("%d/%m/%Y")
                total_cost = self.frames[1].inputs[0].get() + self.frames[1].inputs[1].get() + self.frames[1].inputs[
                    2].get() + self.frames[1].inputs[3].get()
                new_row = (ws.max_row, new_today_date, new_due_date, "£" + str(total_cost), 'Not Paid')

                src_dir = os.getcwd()  # get the current working dir
                print(src_dir)

                # create a dir where we want to copy and rename

                self.dest_dir = expanduser('~/Documents/Invoices')
                self.src_file = os.path.join(src_dir, 'preview_pdf.pdf')
                self.copy_same_file_pass(self.src_file, self.dest_dir)
                #shutil.copy(self.src_file, self.dest_dir)  # copy the file to destination dir

                self.dst_file = os.path.join(self.dest_dir, 'preview_pdf.pdf')
                self.new_dst_file_name = os.path.join(self.dest_dir, 'Invoice-' + str(ws.max_row) + '.pdf')

                os.rename(self.dst_file, self.new_dst_file_name)  # rename
                os.chdir(self.dest_dir)

                ws.append(new_row)
                wb.save(expanduser('~/Documents/Invoices/invoice_list.xlsx'))

                olApp = win32.Dispatch("outlook.application")
                olNS = olApp.GetNameSpace('MAPI')

                MailItem = olApp.CreateItem(0)
                MailItem.Subject = 'Invoice - ' + str(ws.max_row - 1)
                MailItem.BodyFormat = 1
                MailItem.Body = 'Dear ' + self.frames[0].inputs[0].get() + ',\n\nI hope your well. Please see attached invoice number - ' + str(ws.max_row - 1) + ' for the work done at, ' + self.frames[0].inputs[3].get() + ' which is due for payment on ' + new_due_date + '.\n\nAny issue please contact me on my mobile.\n\nKind Regards,\n' #enter users name here
                #MailItem.Sender = #input your email address here
                MailItem.To = self.frames[0].inputs[1].get()
                filename = 'Invoice-' + str(ws.max_row - 1) + '.pdf'
                print(os.path.join(expanduser(r'~\Documents\Invoices'), filename))
                MailItem.Attachments.Add(os.path.join(expanduser(r'~\Documents\Invoices'), filename))

                # To display the mail before sending it
                MailItem.Display()
                MailItem.Save()

                self.frames[3].grid_forget()
                self.frames[0].inputs[0].delete(0, 'end')
                self.frames[0].inputs[1].delete(0, 'end')
                self.frames[0].inputs[2].delete(0, 'end')
                self.frames[0].inputs[3].delete(0, 'end')
                self.frames[1].inputs[0].set(0)
                self.frames[1].inputs[1].set(0)
                self.frames[1].inputs[2].set(0)
                self.frames[1].inputs[3].set(0)
                self.frames[2].textbox.delete(1.0, "end-1c")
                self.dest_dir = None
                self.src_file = None
                self.dst_file = None
                self.new_dst_file_name = None
                self.frames[0].grid(row=0, column=0, padx=10, pady=(10, 0), sticky="nsew")

                #MailItem.Send()
                # app.destroy()

    def previousbutton_callback(self):
        if self.frames[1].winfo_ismapped():
            self.frames[1].grid_forget()
            self.frames[0].grid(row=0, column=0, padx=10, pady=(10, 0), sticky="nsew")
        elif self.frames[2].winfo_ismapped():
            self.frames[2].grid_forget()
            self.frames[1].grid(row=0, column=0, padx=10, pady=(10, 0), sticky="nsew")
        elif self.frames[3].winfo_ismapped():
            self.frames[3].grid_forget()
            self.frames[2].grid(row=0, column=0, padx=10, pady=(10, 0), sticky="nsew")

    def get_page(self, page_class):
        return self.frames[page_class]


app = App()
app.mainloop()
